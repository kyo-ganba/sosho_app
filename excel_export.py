"""
送迎表Excel出力 — Ⅴ番館フォーマット完全再現版

実測した列構造（1-indexed, openpyxl準拠）:
  セレナ:    時刻=B(2), 迎送=C(3), 氏名=D(4), 場所=E(5), 担当=F(6)  車両名=D5(D:E結合)
  ボクシー:  時刻=J(10),迎送=K(11),氏名=L(12),場所=M(13),担当=N(14) 車両名=K5(K:L結合)
  白フリード:時刻=O(15),迎送=P(16),氏名=Q(17),場所=R(18),担当=S(19) 車両名=Q5(Q:R結合)
  銀フリ地下:時刻=U(21),迎送=V(22),氏名=W(23),場所=X(24),担当=Y(25) 車両名=W5(W:X結合)
  銀フリ浅香:時刻=AK(37),迎送=AL(38),氏名=AM(39),場所=AN(40),担当=AO(41) 車両名=AC5(AC:AD結合)
  自力:      列AG(33)=時刻, AH(34)=氏名  ← 迎送なし、名前のみ
  右エリア1: 時刻=AR(44),迎送=AS(45),氏名=AT(46),場所=AU(47),担当=AV(48) ← 銀フリ浅香の第2グループ
  右エリア2: 時刻=AX(50),氏名=AY(51),備考=AZ(52) ← 自力送迎者リスト

色:
  迎えセル(迎送列): ピンク #FF7C80
  送りセル(迎送列): 緑   #00CC99
  児発氏名セル:     黄   #FFFF00
  放デイ氏名セル:   黄   #FFFF00  ← 元は同じ黄色
"""

import io
from datetime import date
from typing import Dict, List, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 色定数（元ファイル実測値）────────────────────────────────
PINK      = "FFFF7C80"   # 迎えセル
GREEN_C   = "FF00CC99"   # 送りセル
YELLOW    = "FFFFFF00"   # 氏名セル（放デイ・共通）
YELLOW_JIS= "FFFFCC00"   # 氏名セル（児発 ← 少し濃い黄で区別）
LIGHT_BLUE= "FFD9E1F2"   # 児発の迎えセル背景（区別用）
ORANGE    = "FFFFD966"   # 児発の送りセル背景（区別用）

THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── 車両グループ列定義（1-indexed）─────────────────────────────
# (時刻, 迎送, 氏名, 場所, 担当, 車両名表示列, 車両名結合終列)
VEHICLE_COLS = {
    "セレナ":    (2,  3,  4,  5,  6,  4,  5),
    "ボクシー":  (10, 11, 12, 13, 14, 11, 12),
    "白フリード":(15, 16, 17, 18, 19, 17, 18),
    "銀フリ地下":(21, 22, 23, 24, 25, 23, 24),
    "銀フリ浅香":(37, 38, 39, 40, 41, 29, 30),  # 車両名はAC(29)
}

# 自力送迎（迎送列なし、時刻+氏名のみ）
JIRIKI_TIME_COL = 33   # AG
JIRIKI_NAME_COL = 34   # AH

# 銀フリ浅香の第2ルート（AR列以降）
EXTRA_TIME_COL   = 44  # AR
EXTRA_MS_COL     = 45  # AS
EXTRA_NAME_COL   = 46  # AT
EXTRA_PLACE_COL  = 47  # AU
EXTRA_DRIVER_COL = 48  # AV

# 自力送迎リスト（AX列以降）
JIRIKI2_TIME_COL = 50  # AX
JIRIKI2_NAME_COL = 51  # AY
JIRIKI2_NOTE_COL = 52  # AZ

# スタッフ名配置列（行2: 各車両グループの氏名列に対応）
STAFF_HEADER_COLS = [4, 11, 17, 23, 31]  # D,K,Q,W,AE

ROW_DATE      = 2
ROW_WORKTIME  = 3
ROW_COUNT     = 4
ROW_VEH_HDR  = 5
ROW_DATA_START= 6


def export_schedule(
    routes: Dict[str, Dict[str, List[Dict]]],
    target_date: date,
    館: str,
    staff_on_duty: Dict,
    jiriki_users: List[Dict] = None,
) -> bytes:
    """
    routes: {"セレナ": {"迎え": [...], "送り": [...]}, ...}
    jiriki_users: [{"time":"10:00","name":"山田太郎","note":"自力"}, ...]
    """
    wb = Workbook()
    ws = wb.active
    weekday_jp = ["月", "火", "水", "木", "金", "土", "日"][target_date.weekday()]
    ws.title = f"{target_date.month}月{target_date.day}日({weekday_jp})"

    _set_col_widths(ws)
    _set_row_heights(ws)
    _write_date_staff(ws, target_date, weekday_jp, 館, staff_on_duty)
    _write_vehicle_headers(ws, routes)
    _write_child_count_formula(ws, routes)
    _write_trips(ws, routes)
    _write_jiriki(ws, jiriki_users or [])

    return _to_bytes(wb)


def _write_date_staff(ws, target_date, weekday_jp, 館, staff_on_duty):
    """行2: 日付・曜日・スタッフ名  行3: 勤務時間"""
    # 日付（D2:F2結合）
    ws.merge_cells(start_row=ROW_DATE, start_column=4, end_row=ROW_DATE, end_column=6)
    _c(ws, ROW_DATE, 4,
       f"{target_date.year}年{target_date.month}月{target_date.day}日",
       bold=True, size=12, ha="center")
    _c(ws, ROW_DATE, 7, f"（{weekday_jp}）", bold=True, size=12, ha="center")

    # スタッフ名と勤務時間（各車両グループ列に配置）
    staff_list = [(n, i) for n, i in staff_on_duty.items() if i.get("on")]
    for idx, (name, info) in enumerate(staff_list):
        if idx >= len(STAFF_HEADER_COLS):
            break
        col = STAFF_HEADER_COLS[idx]
        end_col = col + 1
        ws.merge_cells(start_row=ROW_DATE, start_column=col,
                       end_row=ROW_DATE, end_column=end_col)
        ws.merge_cells(start_row=ROW_WORKTIME, start_column=col,
                       end_row=ROW_WORKTIME, end_column=end_col)
        _c(ws, ROW_DATE,     col, name,                    size=11, ha="center")
        _c(ws, ROW_WORKTIME, col, info.get("work_time","9:30-18:30"), size=9,  ha="center")


def _write_vehicle_headers(ws, routes):
    """行5: 車両名(結合)・運転・添乗・LINE"""
    active_vehicles = list(routes.keys())

    for vehicle, col_def in VEHICLE_COLS.items():
        t_col, ms_col, nm_col, pl_col, dr_col, vn_col, vn_end = col_def

        # 銀フリ浅香の車両名はAC列
        ws.merge_cells(start_row=ROW_VEH_HDR, start_column=vn_col,
                       end_row=ROW_VEH_HDR, end_column=vn_end)
        c = ws.cell(row=ROW_VEH_HDR, column=vn_col)
        c.value = vehicle
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")

        # 運転・LINE ラベル
        _c(ws, ROW_VEH_HDR, dr_col,   "運転", size=9, ha="center")
        _c(ws, ROW_VEH_HDR, dr_col+1, "LINE", size=9, ha="center")

    # 自力ヘッダー
    _c(ws, ROW_VEH_HDR, JIRIKI_NAME_COL, "自力", bold=True, size=11, ha="center")

    # 右エリアヘッダー
    _c(ws, ROW_VEH_HDR, EXTRA_MS_COL,  "迎/送", size=9, ha="center")
    _c(ws, ROW_VEH_HDR, EXTRA_NAME_COL,"氏名",  size=9, ha="center")
    _c(ws, ROW_VEH_HDR, JIRIKI2_NAME_COL, "自力送迎", bold=True, size=10, ha="center")


def _write_child_count_formula(ws, routes):
    """行4: 各車両の児発（(児)を含む）人数カウント"""
    # 氏名列ごとにCOUNTIF数式を設定
    # データ範囲: 行6〜行50
    count_defs = {
        "セレナ":    4,   # D列
        "ボクシー":  12,  # L列
        "白フリード":17,  # Q列
        "銀フリ地下":23,  # W列
        "銀フリ浅香":39,  # AM列
    }
    for vehicle, col in count_defs.items():
        col_letter = get_column_letter(col)
        formula = f'="チャイルド"&COUNTIF({col_letter}6:{col_letter}50,"*(児)")'
        c = ws.cell(row=ROW_COUNT, column=col, value=formula)
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")


def _write_trips(ws, routes):
    """行6〜: 迎え/送りデータを各車両列に書き込む"""
    # 車両ごとに現在の書き込み行を管理
    cur = {v: ROW_DATA_START for v in list(VEHICLE_COLS.keys()) + ["extra"]}

    # 全便を時刻順でフラット化
    all_trips = []
    for vehicle, type_map in routes.items():
        for trip_type, trips in type_map.items():
            for trip in trips:
                all_trips.append((vehicle, trip))
    all_trips.sort(key=lambda x: x[1].get("time", "00:00"))

    for vehicle, trip in all_trips:
        if vehicle not in VEHICLE_COLS:
            continue
        t_col, ms_col, nm_col, pl_col, dr_col, _, _ = VEHICLE_COLS[vehicle]
        r = cur[vehicle]

        _write_trip_row(ws, r, t_col, ms_col, nm_col, pl_col, dr_col, trip)
        cur[vehicle] += 1

    # リーダー欄（全車両の最終行＋2）
    last_row = max(cur.values()) + 2
    _c(ws, last_row, 4,  "リーダー", bold=True, size=10)
    _c(ws, last_row, 5,  "児発：",   size=10)
    _c(ws, last_row, 10, "放デイ：", size=10)


def _write_trip_row(ws, r, t_col, ms_col, nm_col, pl_col, dr_col, trip):
    """1便分のデータを1行に書く"""
    ms      = trip.get("type", "")
    name    = trip.get("name", "")
    place   = trip.get("place", "自宅")
    driver  = trip.get("driver", "")
    t_str   = trip.get("time", "")
    is_jis  = "(児)" in name

    # 迎送の色分け（放デイ/児発で変える）
    if ms == "迎え":
        ms_color   = LIGHT_BLUE if is_jis else PINK
        nm_color   = YELLOW_JIS if is_jis else YELLOW
    else:
        ms_color   = ORANGE     if is_jis else GREEN_C
        nm_color   = YELLOW_JIS if is_jis else YELLOW

    _c(ws, r, t_col,  t_str,                        size=11, ha="center", border=True)
    _c(ws, r, ms_col, "迎" if ms == "迎え" else "送",
       size=11, ha="center", bg=ms_color, border=True)
    _c(ws, r, nm_col, name,   size=11, ha="center", bg=nm_color, border=True)
    _c(ws, r, pl_col, place,  size=11, ha="center", border=True)
    _c(ws, r, dr_col, driver, size=11, ha="center", border=True)


def _write_jiriki(ws, jiriki_users: List[Dict]):
    """自力送迎者リストを右エリア（AX列〜）に書く"""
    if not jiriki_users:
        return
    r = ROW_DATA_START
    for u in jiriki_users:
        _c(ws, r, JIRIKI2_TIME_COL, u.get("time", ""),   size=10, ha="center")
        _c(ws, r, JIRIKI2_NAME_COL, u.get("name", ""),   size=10, ha="left")
        _c(ws, r, JIRIKI2_NOTE_COL, u.get("note", "自力"), size=9, ha="left")
        r += 1


def _set_col_widths(ws):
    """元ファイルの列幅を再現"""
    widths = {
        'A':22.4,'B':7.5,'C':6.0,'D':13.0,'E':9.7,'F':8.7,'G':8.9,'H':13.0,
        'I':7.6,'J':4.9,'K':13.0,'L':10.2,'M':8.9,'N':8.0,'O':7.6,'P':5.0,
        'Q':13.0,'R':9.0,'S':8.9,'T':8.6,'U':7.6,'V':5.0,'W':13.0,'X':9.0,
        'Y':8.9,'Z':8.0,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for i in range(27, 56):
        ws.column_dimensions[get_column_letter(i)].width = 8.5


def _set_row_heights(ws):
    heights = {1:45.75, 2:48.75, 3:41.25, 4:31.5, 5:30.0}
    for r, h in heights.items():
        ws.row_dimensions[r].height = h
    for r in range(6, 55):
        ws.row_dimensions[r].height = 30.0


def _c(ws, row, col, value, bold=False, size=11, ha="center",
       bg=None, border=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size)
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=True)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = BORDER
    return c


def _to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
